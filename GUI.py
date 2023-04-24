import json
from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process_input', methods=['POST'])
def process_input():
    input_text = request.form['input_text']
    # process input_text here
    return ''

@app.route('/process_file', methods=['POST'])
def process_file():
    file = request.files['file']
    workbook = load_workbook(file.filename, read_only=True)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    headers = data.pop(0)
    data_dict = {
        "headers": headers,
        "data": data
    }
    return json.dumps(data_dict)

if __name__ == '__main__':
    app.run(debug=True)
